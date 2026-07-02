import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Respuestas"

    # Definir los encabezados básicos
    headers = [
        "fechaEnvio",
        "nombreCompleto",
        "identificacion",
        "telefono",
        "correo",
        "cargo",
        "institucion",
        "puntajeTotal",
        "nivelDesempeno"
    ]

    # Agregar encabezados por cada uno de los 8 dominios
    for i in range(1, 9):
        headers.append(f"dominio_{i}_puntaje")
        headers.append(f"dominio_{i}_porcentaje")

    # Agregar encabezados para las 43 preguntas individuales
    for i in range(1, 44):
        headers.append(f"pregunta_{i}")

    # Escribir encabezados en la primera fila
    ws.append(headers)

    # Convertir el rango de celdas en una Tabla de Excel (requerido por Power Automate)
    # Por ejemplo, de la celda A1 a la última columna en la fila 2 (dejamos una fila vacía para que tenga estructura de tabla)
    num_cols = len(headers)
    last_col_letter = openpyxl.utils.get_column_letter(num_cols)
    table_range = f"A1:{last_col_letter}2"

    # Rellenar la fila 2 con valores en blanco para evitar problemas
    ws.append([""] * num_cols)

    # Crear tabla de openpyxl
    tab = Table(displayName="TablaRespuestas", ref=table_range)

    # Agregar estilo visual a la tabla
    style = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Ajustar ancho de las columnas para que sea legible
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save("Respuestas_Autoevaluacion.xlsx")
    print("¡Plantilla de Excel creada exitosamente con el nombre: Respuestas_Autoevaluacion.xlsx!")

if __name__ == "__main__":
    create_excel_template()
